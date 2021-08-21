import os

import openpyxl
from PyQt5.QtCore import pyqtSlot

import var
from backend import back


def add_new_paste_1():
    nazwa = var.new_paste_resistor_name
    r = var.new_paste_r
    db = var.db_paste_rezystywne
    dataframe = var.db_pasty_rezystywne_dataframe
    sheet_name = var.db_pasty_rezystywne_sheet

    if type(nazwa) == str and type(r) == float:
        db.update({nazwa: {dataframe[1]: float(r)}})

    try:
        workbook = openpyxl.load_workbook(var.path_pasty)
        sheet = workbook.get_sheet_by_name(sheet_name)
        sheet.append((str(nazwa), r))
        workbook.save(var.path_pasty)

    except FileNotFoundError or FileExistsError:
        init_csv()
        pasty_to_dict()

    print("r:", var.db_paste_rezystywne)


def add_new_paste_2():
    nazwa = var.new_paste_capacitor_name
    db = var.db_paste_dielektryczne
    przenikalnosc = var.new_paste_przenikalnosc
    wytrzymalosc = var.new_paste_voltage
    dataframe = var.db_pasty_dielektryczne_dataframe
    sheet_name = var.db_pasty_dielektryczne_sheet

    print("nazwa", nazwa)
    print("db", db)
    print("przenikalnosc", przenikalnosc)
    print("wytrzymalosc ", wytrzymalosc)
    print("dataframe", dataframe)
    print("sheet_name", sheet_name)

    if type(nazwa) == str and type(przenikalnosc) == float and type(wytrzymalosc) == float:
        db.update({nazwa: {dataframe[1]: float(przenikalnosc),
                           dataframe[2]: float(wytrzymalosc)}})

    try:
        workbook = openpyxl.load_workbook(var.path_pasty)
        sheet = workbook.get_sheet_by_name(sheet_name)
        sheet.append((str(nazwa), przenikalnosc, wytrzymalosc))
        workbook.save(var.path_pasty)

    except FileNotFoundError or FileExistsError:
        init_csv()
        pasty_to_dict()

    print("dielektryk:", var.db_paste_dielektryczne)


def init_csv():
    workbook = openpyxl.Workbook()

    sheet_rezystywne = workbook.create_sheet(var.db_pasty_rezystywne_sheet, 0)
    sheet_izolacyjne = workbook.create_sheet(var.db_pasty_dielektryczne_sheet, 1)
    workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))

    for column in range(1, len(var.db_pasty_rezystywne_dataframe) + 1):
        sheet_rezystywne.cell(row=1, column=column).value = var.db_pasty_rezystywne_dataframe[column - 1]

        row = 2
        for paste in var.db_pasty_rezystywne_default:
            sheet_rezystywne.cell(row=row, column=column).value = paste[column - 1]
            row = row + 1

    for column in range(1, len(var.db_pasty_dielektryczne_dataframe) + 1):
        sheet_izolacyjne.cell(row=1, column=column).value = var.db_pasty_dielektryczne_dataframe[column - 1]

        row = 2
        for paste in var.db_pasty_izolacyjne_default:
            sheet_izolacyjne.cell(row=row, column=column).value = paste[column - 1]
            row = row + 1

    workbook.save(var.path_pasty)


def pasty_to_dict():
    workbook = openpyxl.load_workbook(var.path_pasty)

    sheet_rezystywne = workbook.get_sheet_by_name(var.db_pasty_rezystywne_sheet)
    sheet_dielectric = workbook.get_sheet_by_name(var.db_pasty_dielektryczne_sheet)

    sheet_list = [sheet_rezystywne, var.db_pasty_rezystywne_dataframe, var.db_paste_rezystywne]

    for row in sheet_list[0].iter_rows(min_row=2, min_col=1, max_row=30, max_col=len(sheet_list[1])):
        tmp = []
        for cell in row:
            tmp.append(cell.value)

        nazwa = str(tmp[0])
        r = str(tmp[1])

        if r != "None":
            r = input_to_float(r)

            if type(nazwa) != "" and type(r) == float:
                sheet_list[2].update({nazwa: {sheet_list[1][1]: r}})

    sheet_list = [sheet_dielectric, var.db_pasty_dielektryczne_dataframe, var.db_paste_dielektryczne]

    for row in sheet_list[0].iter_rows(min_row=2, min_col=1, max_row=30, max_col=len(sheet_list[1])):
        tmp = []
        for cell in row:
            tmp.append(cell.value)

        nazwa = str(tmp[0])
        przenikalnosc_min = str(tmp[1])
        przenikalnosc_max = str(tmp[2])
        wytrzymalosc = str(tmp[3])

        if przenikalnosc_min != "None" and przenikalnosc_max != "None" and wytrzymalosc != "None":
            przenikalnosc_min = input_to_float(przenikalnosc_min)
            przenikalnosc_max = input_to_float(przenikalnosc_max)
            wytrzymalosc = input_to_float(wytrzymalosc)

            if type(nazwa) != "" \
                    and type(przenikalnosc_min) == float \
                    and type(przenikalnosc_max) == float \
                    and type(wytrzymalosc) == float:

                sheet_list[2].update({nazwa: {sheet_list[1][1]: przenikalnosc_min,
                                              sheet_list[1][2]: przenikalnosc_max,
                                              sheet_list[1][3]: wytrzymalosc}})

    print(var.db_paste_dielektryczne)
    print(var.db_paste_rezystywne)


def init():
    if not os.path.isfile(var.path_pasty):
        init_csv()
    pasty_to_dict()


class static_function(back.application):

    @pyqtSlot()
    def exit(self):
        print("exit")
        exit(1)

    @pyqtSlot()
    def add_new_paste_1(self):
        add_new_paste_1()

    @pyqtSlot()
    def add_new_paste_2(self):
        add_new_paste_2()


def input_to_float(inside):
    try:
        inside = inside.replace(",", ".")
        value = float(inside)
        return value
    except AttributeError:
        value = float(inside)
        return value
    except TypeError:
        return 0
    except ValueError:
        return 0
